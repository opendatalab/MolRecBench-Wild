#!/usr/bin/env python3
"""Evaluate every supported prediction file and save the results to Excel.

The script recursively discovers SMILES, Graph, and simplified Graph
predictions below ``results/``.  Every prediction is evaluated through the
self-contained evaluators in this package.  Unsupported files and per-file
errors are recorded in the workbook instead of aborting the whole batch.
"""

from __future__ import annotations

import argparse
import fnmatch
import json
import os
import platform
import re
import stat
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence


if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GT = REPO_ROOT / "dataset" / "annotation.jsonl"
DEFAULT_RESULTS_ROOT = REPO_ROOT / "results"
DEFAULT_OUTPUT = DEFAULT_RESULTS_ROOT / "evaluation_results.xlsx"

TRACK_LABELS = {
    "smiles": "SMILES",
    "s_graph": "S-Graph",
    "graph": "Graph",
}

PAPER_SUBSETS = ("Full", "A", "B", "C")
DISPLAY_METHOD_NAMES = {
    "OCSU": "OCSU",
    "DECIMERv2.2": "DECIMER v2.2",
    "MolGrapher": "MolGrapher",
    "MolNexTR": "MolNexTR",
    "MolScribe": "MolScribe",
    "ChemDFM-X": "ChemDFM-X",
    "ChemVLM": "ChemVLM",
    "Logics-Parsing": "Logic-Parsing",
    "Mathpix": "Mathpix",
    "InternVL3_5": "InternVL3.5",
    "GLM-4.5V": "GLM-4.5V",
    "Intern-S1": "Intern-S1",
    "Seed1.6-Thinking": "Seed1.6-Thinking",
    "Claude-opus-4-8": "Claude-opus-4-8",
    "Gemini-3.5-flash-thinking": "Gemini-3.5-flash-thinking",
    "GPT-5.6-sol": "GPT-5.6-Sol",
    "GTR-VL-1.4.13": "MinerU.Chem\n(GTR-VL-1.4.13)",
}

PAPER_METHOD_ORDER = (
    "OCSU",
    "DECIMERv2.2",
    "MolGrapher",
    "MolNexTR",
    "MolScribe",
    "ChemDFM-X",
    "ChemVLM",
    "Logics-Parsing",
    "Mathpix",
    "InternVL3_5",
    "GLM-4.5V",
    "Intern-S1",
    "Seed1.6-Thinking",
    "Claude-opus-4-8",
    "Gemini-3.5-flash-thinking",
    "GPT-5.6-sol",
    "GTR-VL-1.4.13",
)

EVALUATION_COLUMNS = (
    ("category", "Category"),
    ("model", "Model"),
    ("variant", "Variant"),
    ("result_name", "Result Name"),
    ("track", "Track"),
    ("status", "Status"),
    ("accuracy", "Accuracy"),
    ("accuracy_over_all_gt", "Accuracy / All GT"),
    ("correct_records", "Correct Records"),
    ("scored_records", "Scored Records"),
    ("total_gt_records", "Evaluated GT Records"),
    ("prediction_records", "Prediction Records"),
    ("coverage", "GT ID Coverage"),
    ("matched_gt_ids", "Matched GT IDs"),
    ("missing_gt_ids", "Missing GT IDs"),
    ("extra_prediction_ids", "Extra Prediction IDs"),
    ("prediction_unique_ids", "Prediction Unique IDs"),
    ("duplicate_prediction_ids", "Duplicate Prediction IDs"),
    ("filtered_records", "Filtered / Invalid GT"),
    (
        "gt_load_success_records",
        "GT Conversion / Load Success",
    ),
    (
        "prediction_load_success_records",
        "Prediction Load Success",
    ),
    ("read_error_count", "Read Errors"),
    ("index_error_count", "Index Errors"),
    ("elapsed_seconds", "Elapsed Seconds"),
    ("prediction_file", "Prediction File"),
    ("notes", "Notes"),
    ("error", "Error"),
)

SUMMARY_COLUMNS = (
    ("category", "Category"),
    ("model", "Model"),
    ("variant", "Variant"),
    ("smiles_accuracy", "SMILES Accuracy"),
    ("smiles_status", "SMILES Status"),
    ("s_graph_accuracy", "S-Graph Accuracy"),
    ("s_graph_status", "S-Graph Status"),
    ("graph_accuracy", "Graph Accuracy"),
    ("graph_status", "Graph Status"),
    ("completed_tracks", "Completed Tracks"),
    ("result_files", "Result Files"),
)

SKIPPED_COLUMNS = (
    ("category", "Category"),
    ("model", "Model"),
    ("file", "File"),
    ("reason", "Reason"),
)


@dataclass(frozen=True)
class EvaluationJob:
    """One prediction file and its inferred metric."""

    path: Path
    relative_path: str
    category: str
    model: str
    result_name: str
    variant: str
    track: str


@dataclass(frozen=True)
class IdStats:
    """ID information collected without loading chemistry dependencies."""

    record_count: int
    ordered_ids: tuple[str, ...]
    unique_ids: frozenset[str]
    duplicate_count: int
    read_error_count: int
    missing_id_count: int


def detect_track(path: Path) -> str | None:
    """Infer the evaluation track from a result filename."""

    if path.suffix.lower() != ".jsonl":
        return None
    stem = path.stem.lower()
    if "graph_simple" in stem or "s_graph" in stem:
        return "s_graph"
    if re.search(r"(?:^|[_-])graph$", stem):
        return "graph"
    if "smiles" in stem:
        return "smiles"
    if stem.endswith("_carbon") or stem.endswith("-carbon"):
        return "graph"
    return None


def is_gtr_carbon_prediction(path: Path, model: str) -> bool:
    """Return whether the file is the MinerU.Chem Carbon prediction file."""

    return (
        model == "GTR-VL-1.4.13"
        and path.name == "GTR-VL-1.4.13.jsonl"
    )


def result_variant(path: Path) -> str:
    """Remove a terminal metric suffix while retaining the run identity."""

    return re.sub(
        r"(?i)(?:_graph_simple|-graph-simple|_s_graph|-s-graph|"
        r"_smiles|-smiles|_graph|-graph|_carbon|-carbon)$",
        "",
        path.stem,
    )


def _matches_filters(
    relative_path: str,
    includes: Sequence[str],
    excludes: Sequence[str],
) -> bool:
    included = not includes or any(fnmatch.fnmatch(relative_path, pattern) for pattern in includes)
    excluded = any(fnmatch.fnmatch(relative_path, pattern) for pattern in excludes)
    return included and not excluded


def _path_identity(path: Path, root: Path) -> tuple[str, str, str]:
    relative = path.relative_to(root).as_posix()
    parts = Path(relative).parts
    category = parts[0] if len(parts) > 1 else "root"
    model = parts[-2] if len(parts) > 1 else "root"
    return relative, category, model


def discover_jobs(
    results_root: Path,
    *,
    includes: Sequence[str] = (),
    excludes: Sequence[str] = (),
) -> tuple[list[EvaluationJob], list[dict[str, str]], dict[str, int]]:
    """Discover supported JSONL files and auditable skipped files."""

    jobs: list[EvaluationJob] = []
    skipped: list[dict[str, str]] = []
    jsonl_count = 0
    csv_count = 0

    for path in sorted(
        (item for item in results_root.rglob("*") if item.is_file()),
        key=lambda item: item.relative_to(results_root).as_posix(),
    ):
        relative, category, model = _path_identity(path, results_root)
        if not _matches_filters(relative, includes, excludes):
            continue
        suffix = path.suffix.lower()
        if suffix == ".csv":
            csv_count += 1
            skipped.append(
                {
                    "category": category,
                    "model": model,
                    "file": relative,
                    "reason": ("CSV intermediate file; the JSONL result is evaluated to avoid duplicate scoring."),
                }
            )
            continue
        if suffix != ".jsonl":
            continue
        jsonl_count += 1
        track = detect_track(path)
        if track is None and is_gtr_carbon_prediction(path, model):
            track = "graph"
        if track is None:
            skipped.append(
                {
                    "category": category,
                    "model": model,
                    "file": relative,
                    "reason": ("Cannot infer SMILES, S-Graph, or Graph track from the filename."),
                }
            )
            continue
        jobs.append(
            EvaluationJob(
                path=path,
                relative_path=relative,
                category=category,
                model=model,
                result_name=path.stem,
                variant=result_variant(path),
                track=track,
            )
        )
        if (
            model == "GTR-VL-1.4.13"
            and track == "graph"
            and (
                path.stem.lower().endswith("_carbon")
                or is_gtr_carbon_prediction(path, model)
            )
        ):
            jobs.append(
                EvaluationJob(
                    path=path,
                    relative_path=relative,
                    category=category,
                    model=model,
                    result_name=f"{path.stem}_smiles",
                    variant=result_variant(path),
                    track="smiles",
                )
            )

    counts = {
        "jsonl_files": jsonl_count,
        "csv_files": csv_count,
        "identified_files": len(jobs),
        "skipped_files": len(skipped),
    }
    for track in TRACK_LABELS:
        counts[f"identified_{track}"] = sum(job.track == track for job in jobs)
    return jobs, skipped, counts


def collect_id_stats(path: Path) -> IdStats:
    """Read valid string IDs from a JSONL file for coverage reporting."""

    record_count = 0
    ordered_ids: list[str] = []
    unique_ids: set[str] = set()
    duplicate_count = 0
    read_error_count = 0
    missing_id_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            record_count += 1
            try:
                record = json.loads(line)
            except (json.JSONDecodeError, UnicodeDecodeError):
                read_error_count += 1
                continue
            if not isinstance(record, dict):
                read_error_count += 1
                continue
            record_id = record.get("id")
            if not isinstance(record_id, str) or not record_id:
                missing_id_count += 1
                continue
            ordered_ids.append(record_id)
            if record_id in unique_ids:
                duplicate_count += 1
            else:
                unique_ids.add(record_id)
    return IdStats(
        record_count=record_count,
        ordered_ids=tuple(ordered_ids),
        unique_ids=frozenset(unique_ids),
        duplicate_count=duplicate_count,
        read_error_count=read_error_count,
        missing_id_count=missing_id_count,
    )


def _run_smiles_evaluation(
    gt_path: Path,
    pred_path: Path,
    *,
    ignore_cistrans: bool,
    limit: int | None,
    allow_coverage_mismatch: bool,
) -> dict[str, Any]:
    from evaluate.smiles_metric import evaluate

    return evaluate(
        gt_path=gt_path,
        pred_path=pred_path,
        output_path=None,
        summary_path=None,
        missing_abbreviations_path=None,
        ignore_cistrans=ignore_cistrans,
        limit=limit,
        output_csv_path=None,
        allow_coverage_mismatch=allow_coverage_mismatch,
    )


def _run_carbon_smiles_evaluation(
    gt_path: Path,
    pred_path: Path,
    *,
    ignore_cistrans: bool,
    limit: int | None,
    allow_coverage_mismatch: bool,
) -> dict[str, Any]:
    """Evaluate graph predictions as SMILES after benchmark filtering."""

    from evaluate.Evaluator import Evaluator
    from evaluate.smiles_metric import (
        ANNOTATION_METADATA_FIELDS,
        filter_reasons,
        normalize_smiles,
        read_jsonl,
        prediction_coverage_errors,
    )
    from evaluate.utils import carbon_to_smiles

    gt_records, gt_read_errors = read_jsonl(gt_path)
    pred_records, pred_read_errors = read_jsonl(pred_path)
    full_gt_records = gt_records
    if limit is not None:
        gt_records = gt_records[:limit]
    coverage_errors = prediction_coverage_errors(
        full_gt_records,
        gt_records,
        pred_records,
        prediction_read_errors=pred_read_errors,
    )
    if coverage_errors and not allow_coverage_mismatch:
        raise ValueError(
            "Prediction coverage validation failed: "
            + "; ".join(coverage_errors)
        )

    filtered_gt: list[dict[str, Any]] = []
    for record in gt_records:
        if filter_reasons(record, ANNOTATION_METADATA_FIELDS):
            continue
        try:
            gt_raw = carbon_to_smiles(record)
            _, _, _, expansion_ok, canonical_ok, missing = normalize_smiles(
                gt_raw, {}, ignore_cistrans
            )
        except Exception:
            continue
        if expansion_ok and canonical_ok and not missing:
            filtered_gt.append(record)

    evaluator = Evaluator(
        gt_list=filtered_gt,
        pred_list=pred_records,
    )
    success, correct = evaluator.evaluate_smiles(
        expand=True,
        kekule=True,
        ignore_cistrans=ignore_cistrans,
    )
    subset_names = {
        "A": "A",
        "B": "B",
        "C": "C",
    }

    def metric(records: Sequence[dict[str, Any]]) -> dict[str, Any]:
        record_ids = {record["id"] for record in records}
        scored = sum(bool(evaluator.eval_info[record_id].get("smiles_gt")) for record_id in record_ids)
        subset_correct = sum(bool(evaluator.eval_info[record_id].get("smiles_eval")) for record_id in record_ids)
        return {
            "total_gt_records": len(records),
            "scored_records": scored,
            "correct_records": subset_correct,
            "accuracy": subset_correct / scored if scored else 0.0,
        }

    subset_metrics = {"Full": metric(filtered_gt)}
    for display_name, subset_name in subset_names.items():
        records = [record for record in filtered_gt if record.get("evaluation_subset") == subset_name]
        subset_metrics[display_name] = {
            "subset": subset_name,
            **metric(records),
        }
    return {
        "evaluation_method": "Evaluator.evaluate_smiles",
        "ignore_cistrans": ignore_cistrans,
        "total_gt_records": len(gt_records),
        "prediction_records": len(pred_records),
        "included_records": success,
        "filtered_records": len(gt_records) - success,
        "gt_conversion_success_records": success,
        "correct_records": correct,
        "accuracy_on_included_records": correct / success if success else 0.0,
        "accuracy_over_all_gt_records": (correct / len(gt_records) if gt_records else 0.0),
        "gt_read_errors": gt_read_errors,
        "prediction_read_errors": pred_read_errors,
        "index_errors": evaluator.index_errors,
        "coverage_errors": coverage_errors,
        "subset_metrics": subset_metrics,
    }


def _run_graph_evaluation(
    gt_path: Path,
    pred_path: Path,
    *,
    metric: str,
    limit: int | None,
    timeout_seconds: float | None,
    allow_coverage_mismatch: bool,
) -> dict[str, Any]:
    from evaluate.Evaluator import run_graph_evaluation

    return run_graph_evaluation(
        gt_path=gt_path,
        pred_path=pred_path,
        metric=metric,
        split_path=None,
        output_path=None,
        summary_path=None,
        limit=limit,
        timeout_seconds=timeout_seconds,
        allow_coverage_mismatch=allow_coverage_mismatch,
    )


def _empty_result(job: EvaluationJob) -> dict[str, Any]:
    return {
        "category": job.category,
        "model": job.model,
        "variant": job.variant,
        "result_name": job.result_name,
        "track": TRACK_LABELS[job.track],
        "status": "error",
        "accuracy": None,
        "accuracy_over_all_gt": None,
        "correct_records": None,
        "scored_records": None,
        "total_gt_records": None,
        "prediction_records": None,
        "coverage": None,
        "matched_gt_ids": None,
        "missing_gt_ids": None,
        "extra_prediction_ids": None,
        "prediction_unique_ids": None,
        "duplicate_prediction_ids": None,
        "filtered_records": None,
        "gt_load_success_records": None,
        "prediction_load_success_records": None,
        "read_error_count": None,
        "index_error_count": None,
        "elapsed_seconds": None,
        "prediction_file": job.relative_path,
        "notes": "",
        "error": "",
        "subset_metrics": {},
    }


def evaluate_job(
    job: EvaluationJob,
    *,
    gt_path: Path,
    full_gt_ids: frozenset[str],
    evaluated_gt_ids: frozenset[str],
    ignore_cistrans: bool = True,
    limit: int | None = None,
    timeout_seconds: float | None = 5,
    allow_coverage_mismatch: bool = False,
) -> dict[str, Any]:
    """Evaluate one file and normalize the metric-specific summary."""

    started = time.perf_counter()
    row = _empty_result(job)
    prediction_stats: IdStats | None = None
    try:
        prediction_stats = collect_id_stats(job.path)
        matched = len(evaluated_gt_ids & prediction_stats.unique_ids)
        missing = len(evaluated_gt_ids - prediction_stats.unique_ids)
        extra = len(prediction_stats.unique_ids - full_gt_ids)
        row.update(
            {
                "coverage": (matched / len(evaluated_gt_ids) if evaluated_gt_ids else 0.0),
                "matched_gt_ids": matched,
                "missing_gt_ids": missing,
                "extra_prediction_ids": extra,
                "prediction_unique_ids": len(prediction_stats.unique_ids),
                "duplicate_prediction_ids": (prediction_stats.duplicate_count),
            }
        )

        coverage_issues: list[str] = []
        if missing:
            coverage_issues.append(f"{missing} evaluated GT IDs have no prediction")
        if extra:
            coverage_issues.append(f"{extra} prediction IDs are outside the full GT")
        if prediction_stats.duplicate_count:
            coverage_issues.append(
                f"{prediction_stats.duplicate_count} duplicate prediction IDs"
            )
        if prediction_stats.read_error_count:
            coverage_issues.append(
                f"{prediction_stats.read_error_count} prediction read errors"
            )
        if prediction_stats.missing_id_count:
            coverage_issues.append(
                f"{prediction_stats.missing_id_count} prediction rows have no string ID"
            )
        if coverage_issues and not allow_coverage_mismatch:
            raise ValueError(
                "Prediction coverage validation failed: "
                + "; ".join(coverage_issues)
            )

        if job.track == "smiles":
            if job.model == "GTR-VL-1.4.13":
                summary = _run_carbon_smiles_evaluation(
                    gt_path,
                    job.path,
                    ignore_cistrans=ignore_cistrans,
                    limit=limit,
                    allow_coverage_mismatch=allow_coverage_mismatch,
                )
            else:
                summary = _run_smiles_evaluation(
                    gt_path,
                    job.path,
                    ignore_cistrans=ignore_cistrans,
                    limit=limit,
                    allow_coverage_mismatch=allow_coverage_mismatch,
                )
            read_errors = len(summary.get("gt_read_errors", [])) + len(summary.get("prediction_read_errors", []))
            index_errors = len(summary.get("index_errors", []))
            row.update(
                {
                    "accuracy": summary["accuracy_on_included_records"],
                    "accuracy_over_all_gt": summary["accuracy_over_all_gt_records"],
                    "correct_records": summary["correct_records"],
                    "scored_records": summary["included_records"],
                    "total_gt_records": summary["total_gt_records"],
                    "prediction_records": summary["prediction_records"],
                    "filtered_records": summary["filtered_records"],
                    "gt_load_success_records": summary["gt_conversion_success_records"],
                    "prediction_load_success_records": None,
                    "read_error_count": read_errors,
                    "index_error_count": index_errors,
                    "subset_metrics": summary.get("subset_metrics", {}),
                }
            )
        else:
            summary = _run_graph_evaluation(
                gt_path,
                job.path,
                metric=job.track,
                limit=limit,
                timeout_seconds=timeout_seconds,
                allow_coverage_mismatch=allow_coverage_mismatch,
            )
            index_errors = len(summary.get("index_errors", []))
            read_errors = 0
            row.update(
                {
                    "accuracy": summary["accuracy_on_valid_gt_records"],
                    "accuracy_over_all_gt": summary["accuracy_over_all_gt_records"],
                    "correct_records": summary["correct_records"],
                    "scored_records": summary["valid_gt_records"],
                    "total_gt_records": summary["total_gt_records"],
                    "prediction_records": summary["prediction_records"],
                    "filtered_records": (summary["total_gt_records"] - summary["valid_gt_records"]),
                    "gt_load_success_records": summary["gt_load_success_records"],
                    "prediction_load_success_records": summary["prediction_load_success_records"],
                    "read_error_count": read_errors,
                    "index_error_count": index_errors,
                    "subset_metrics": summary.get("subset_metrics", {}),
                }
            )

        notes: list[str] = []
        if missing:
            notes.append(f"{missing} evaluated GT IDs have no prediction")
        if extra:
            notes.append(f"{extra} prediction IDs are outside the full GT")
        if prediction_stats.duplicate_count:
            notes.append(f"{prediction_stats.duplicate_count} duplicate prediction IDs")
        if prediction_stats.read_error_count:
            notes.append(f"{prediction_stats.read_error_count} prediction read errors")
        if prediction_stats.missing_id_count:
            notes.append(f"{prediction_stats.missing_id_count} prediction rows have no string ID")
        if read_errors:
            notes.append(f"evaluator reported {read_errors} read errors")
        if index_errors:
            notes.append(f"evaluator reported {index_errors} index errors")
        pred_load = row["prediction_load_success_records"]
        if pred_load is not None and pred_load < matched:
            notes.append(f"{matched - pred_load} matched predictions failed Carbon loading")
        gt_load = row["gt_load_success_records"]
        total_gt = row["total_gt_records"]
        if job.track != "smiles" and gt_load is not None and total_gt is not None and gt_load < total_gt:
            notes.append(f"{total_gt - gt_load} GT records failed loading")
        row["notes"] = "; ".join(notes)
        invalid_graph_file = job.track != "smiles" and matched > 0 and pred_load == 0
        invalid_gt_file = job.track != "smiles" and bool(total_gt) and gt_load == 0
        if invalid_graph_file:
            row["error"] = "No matched prediction could be loaded as a Carbon graph."
        elif invalid_gt_file:
            row["error"] = "No GT record could be loaded as a Carbon graph."
        if invalid_graph_file or invalid_gt_file:
            row["status"] = "error"
        else:
            row["status"] = "warning" if notes else "success"
    except Exception as exc:
        row["status"] = "error"
        row["error"] = f"{type(exc).__name__}: {exc}"
        if prediction_stats is not None:
            row["prediction_records"] = prediction_stats.record_count
            row["read_error_count"] = prediction_stats.read_error_count
    finally:
        row["elapsed_seconds"] = time.perf_counter() - started
    return row


def build_model_summary(
    result_rows: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build a lossless-enough wide view without merging run variants."""

    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in result_rows:
        base_key = (row["category"], row["model"], row["variant"])
        key = base_key
        metric_key = {
            "SMILES": "smiles",
            "S-Graph": "s_graph",
            "Graph": "graph",
        }[row["track"]]
        serial = 2
        while key in groups and groups[key].get(f"{metric_key}_status") is not None:
            key = (
                base_key[0],
                base_key[1],
                f"{base_key[2]} [{serial}]",
            )
            serial += 1
        group = groups.setdefault(
            key,
            {
                "category": key[0],
                "model": key[1],
                "variant": key[2],
                "smiles_accuracy": None,
                "smiles_status": None,
                "s_graph_accuracy": None,
                "s_graph_status": None,
                "graph_accuracy": None,
                "graph_status": None,
                "completed_tracks": 0,
                "result_files": [],
            },
        )
        group[f"{metric_key}_accuracy"] = row["accuracy"]
        group[f"{metric_key}_status"] = row["status"]
        group["completed_tracks"] += 1
        group["result_files"].append(row["prediction_file"])

    summary_rows = []
    for key in sorted(groups):
        row = groups[key].copy()
        row["result_files"] = "\n".join(row["result_files"])
        summary_rows.append(row)
    return summary_rows


def build_paper_summary(
    result_rows: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build the screenshot-style Method x Subset x Metric result matrix."""

    by_model: dict[str, list[dict[str, Any]]] = {}
    for row in result_rows:
        by_model.setdefault(row["model"], []).append(row)

    ordered_models = [model for model in PAPER_METHOD_ORDER if model in by_model]
    ordered_models.extend(sorted(set(by_model) - set(ordered_models)))
    paper_rows: list[dict[str, Any]] = []
    for model in ordered_models:
        candidates = by_model[model]
        result: dict[str, Any] = {
            "model": model,
            "method": DISPLAY_METHOD_NAMES.get(model, model),
            "category": candidates[0]["category"],
        }
        # Prefer the full Graph track. Expert graph systems in this benchmark
        # only provide S-Graph predictions, which are intentionally shown as '-'.
        for metric_name, track_name in (("smiles", "SMILES"), ("graph", "Graph")):
            rows = [row for row in candidates if row["track"] == track_name]
            row = rows[0] if rows else None
            for subset in PAPER_SUBSETS:
                value = None
                if row is not None and row["status"] != "error":
                    metrics = row.get("subset_metrics") or {}
                    if subset in metrics:
                        value = metrics[subset].get("accuracy")
                    elif subset == "Full":
                        value = row.get("accuracy")
                result[f"{subset.lower()}_{metric_name}"] = value
        paper_rows.append(result)
    return paper_rows


def _write_paper_summary(worksheet: Any, result_rows: Sequence[dict[str, Any]]) -> None:
    """Write a publication-style table matching the reference screenshot."""

    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    rows = build_paper_summary(result_rows)
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B3"

    # Keep only the table itself.  The former caption occupied rows 1--2 and
    # the blank spacer row 3; callers requested that those rows be omitted.
    worksheet.merge_cells("A1:A2")
    worksheet["A1"] = "Method"
    for index, subset in enumerate(PAPER_SUBSETS):
        start = 2 + index * 2
        worksheet.merge_cells(
            start_row=1,
            start_column=start,
            end_row=1,
            end_column=start + 1,
        )
        worksheet.cell(1, start, subset)
        worksheet.cell(2, start, "SMILES")
        worksheet.cell(2, start + 1, "Graph")

    for cell in worksheet[1] + worksheet[2]:
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 22
    worksheet.row_dimensions[2].height = 22

    previous_group: str | None = None
    data_start = 3
    for row_number, row in enumerate(rows, start=data_start):
        group = row["category"]
        worksheet.cell(row_number, 1, row["method"])
        worksheet.cell(row_number, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for subset_index, subset in enumerate(PAPER_SUBSETS):
            for metric_index, metric in enumerate(("smiles", "graph")):
                column = 2 + subset_index * 2 + metric_index
                value = row.get(f"{subset.lower()}_{metric}")
                cell = worksheet.cell(row_number, column)
                if value is None:
                    cell.value = "-"
                else:
                    cell.value = round(float(value) * 100, 2)
                    cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center", vertical="center")
        is_gtr = row["model"] == "GTR-VL-1.4.13"
        for cell in worksheet[row_number]:
            cell.font = Font(name="Times New Roman", size=12, bold=is_gtr)
        if previous_group is not None and group != previous_group:
            for column in range(1, 10):
                cell = worksheet.cell(row_number, column)
                cell.border = Border(top=thin)
        previous_group = group
        worksheet.row_dimensions[row_number].height = 32 if is_gtr else 21

    last_row = worksheet.max_row
    # Bold the best available value and underline the runner-up per column.
    for column in range(2, 10):
        numeric_cells = [worksheet.cell(row_number, column) for row_number in range(data_start, last_row + 1) if isinstance(worksheet.cell(row_number, column).value, float)]
        distinct_values = sorted({cell.value for cell in numeric_cells}, reverse=True)
        for cell in numeric_cells:
            if distinct_values and cell.value == distinct_values[0]:
                cell.font = Font(name="Times New Roman", size=12, bold=True)
            elif len(distinct_values) > 1 and cell.value == distinct_values[1]:
                cell.font = Font(name="Times New Roman", size=12, underline="single")

    for column in range(1, 10):
        worksheet.cell(1, column).border = Border(top=medium)
        bottom_cell = worksheet.cell(last_row, column)
        bottom_cell.border = Border(
            top=bottom_cell.border.top,
            bottom=medium,
        )
    for column in (1, 3, 5, 7, 9):
        for row_number in range(1, last_row + 1):
            cell = worksheet.cell(row_number, column)
            cell.border = Border(
                left=cell.border.left,
                right=thin if column != 9 else cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    worksheet.column_dimensions["A"].width = 31
    for column in range(2, 10):
        worksheet.column_dimensions[get_column_letter(column)].width = 12


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set, frozenset)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_table(
    worksheet: Any,
    columns: Sequence[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
    *,
    percentage_keys: frozenset[str] = frozenset(),
    status_keys: frozenset[str] = frozenset(),
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    status_fills = {
        "success": PatternFill("solid", fgColor="C6EFCE"),
        "warning": PatternFill("solid", fgColor="FFEB9C"),
        "error": PatternFill("solid", fgColor="FFC7CE"),
    }

    worksheet.append([heading for _, heading in columns])
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 24

    materialized_rows = list(rows)
    for row in materialized_rows:
        worksheet.append([_excel_value(row.get(key)) for key, _ in columns])

    for column_number, (key, _) in enumerate(columns, start=1):
        if key in percentage_keys:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row_number, column_number).number_format = "0.0000%"
        elif key == "elapsed_seconds":
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row_number, column_number).number_format = "0.000"
        if key in status_keys:
            for row_number in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_number, column_number)
                fill = status_fills.get(str(cell.value).lower())
                if fill is not None:
                    cell.fill = fill

        values = [str(worksheet.cell(1, column_number).value)]
        values.extend(str(worksheet.cell(row_number, column_number).value or "") for row_number in range(2, worksheet.max_row + 1))
        width = min(max(len(value) for value in values) + 2, 70)
        worksheet.column_dimensions[get_column_letter(column_number)].width = max(width, 10)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(
    output_path: Path,
    result_rows: Sequence[dict[str, Any]],
    skipped_rows: Sequence[dict[str, str]],
    run_info: dict[str, Any],
) -> None:
    """Atomically write the paper table and reproducibility details."""

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write Excel; install the project requirements with `pip install -r requirements.txt`.") from exc

    workbook = Workbook()
    paper_sheet = workbook.active
    paper_sheet.title = "Paper Table"
    _write_paper_summary(paper_sheet, result_rows)

    details_sheet = workbook.create_sheet("Details")
    _write_table(
        details_sheet,
        EVALUATION_COLUMNS,
        result_rows,
        percentage_keys=frozenset(
            {"accuracy", "accuracy_over_all_gt", "coverage"}
        ),
        status_keys=frozenset({"status"}),
    )

    model_summary_sheet = workbook.create_sheet("Model Summary")
    _write_table(
        model_summary_sheet,
        SUMMARY_COLUMNS,
        build_model_summary(result_rows),
        percentage_keys=frozenset(
            {"smiles_accuracy", "s_graph_accuracy", "graph_accuracy"}
        ),
        status_keys=frozenset(
            {"smiles_status", "s_graph_status", "graph_status"}
        ),
    )

    skipped_sheet = workbook.create_sheet("Skipped")
    _write_table(skipped_sheet, SKIPPED_COLUMNS, skipped_rows)

    run_info_sheet = workbook.create_sheet("Run Info")
    _write_table(
        run_info_sheet,
        (("key", "Setting"), ("value", "Value")),
        ({"key": key, "value": value} for key, value in run_info.items()),
    )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_mode = stat.S_IMODE(output_path.stat().st_mode) if output_path.exists() else 0o644
    with tempfile.NamedTemporaryFile(
        dir=output_path.parent,
        prefix=f".{output_path.stem}.",
        suffix=".xlsx",
        delete=False,
    ) as temporary_file:
        temporary_path = Path(temporary_file.name)
    try:
        workbook.save(temporary_path)
        os.chmod(temporary_path, output_mode)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def _runtime_version(module_name: str) -> str:
    """Return the version of the module that this interpreter imports."""

    try:
        module = __import__(module_name)
    except ImportError:
        return "not installed"
    return str(getattr(module, "__version__", "unknown"))


def _run_info(
    *,
    args: argparse.Namespace,
    discovery_counts: dict[str, int],
    selected_jobs: int,
    result_rows: Sequence[dict[str, Any]],
    skipped_rows: Sequence[dict[str, str]],
    state: str,
    started_at: str,
) -> dict[str, Any]:
    status_counts = {status: sum(row["status"] == status for row in result_rows) for status in ("success", "warning", "error")}
    return {
        "state": state,
        "started_at": started_at,
        "last_updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "ground_truth": str(args.gt_path.resolve()),
        "results_root": str(args.results_root.resolve()),
        "output": str(args.output.resolve()),
        "discovered_jsonl_files": discovery_counts["jsonl_files"],
        "discovered_csv_files": discovery_counts["csv_files"],
        "identified_files": discovery_counts["identified_files"],
        "identified_smiles": discovery_counts["identified_smiles"],
        "identified_s_graph": discovery_counts["identified_s_graph"],
        "identified_graph": discovery_counts["identified_graph"],
        "selected_jobs": selected_jobs,
        "completed_jobs": len(result_rows),
        "success_jobs": status_counts["success"],
        "warning_jobs": status_counts["warning"],
        "error_jobs": status_counts["error"],
        "skipped_files": len(skipped_rows),
        "gt_limit": args.limit if args.limit is not None else "all",
        "graph_timeout_seconds": (args.timeout if args.timeout else "disabled"),
        "ignore_cistrans": args.ignore_cistrans,
        "allow_coverage_mismatch": args.allow_coverage_mismatch,
        "include_patterns": args.include or ["*"],
        "exclude_patterns": args.exclude,
        "checkpoint_every": args.checkpoint_every,
        "python": platform.python_version(),
        "rdkit": _runtime_version("rdkit"),
        "openpyxl": _runtime_version("openpyxl"),
        "paper_table_subsets": "Full, A, B, C",
        "subset_A": "A",
        "subset_B": "B",
        "subset_C": "C",
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=("Recursively evaluate all SMILES, Graph, and S-Graph JSONL files below results/ and save a consolidated Excel workbook."))
    parser.add_argument(
        "--gt",
        "--gt-path",
        "--gt_path",
        dest="gt_path",
        type=Path,
        default=DEFAULT_GT,
        help=f"Ground-truth Carbon JSONL (default: {DEFAULT_GT}).",
    )
    parser.add_argument(
        "--results-root",
        "--results-dir",
        "--results_dir",
        dest="results_root",
        type=Path,
        default=DEFAULT_RESULTS_ROOT,
        help=f"Directory scanned recursively (default: {DEFAULT_RESULTS_ROOT}).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx file (default: {DEFAULT_OUTPUT}).",
    )
    parser.add_argument(
        "--include",
        action="append",
        default=[],
        metavar="GLOB",
        help=("Only evaluate relative paths matching this glob; repeatable."),
    )
    parser.add_argument(
        "--exclude",
        action="append",
        default=[],
        metavar="GLOB",
        help="Skip relative paths matching this glob; repeatable.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Evaluate only the first N GT records (for smoke tests).",
    )
    parser.add_argument(
        "--max-files",
        type=int,
        default=None,
        help="Evaluate only the first N discovered files (for smoke tests).",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=5,
        help="Per-record Graph isomorphism timeout; use 0 to disable.",
    )
    cistrans_group = parser.add_mutually_exclusive_group()
    cistrans_group.add_argument(
        "--ignore-cistrans",
        dest="ignore_cistrans",
        action="store_true",
        default=True,
        help=(
            "Ignore double-bond cis/trans markers but still compare atom "
            "chirality (default; matches the benchmark protocol)."
        ),
    )
    cistrans_group.add_argument(
        "--preserve-cistrans",
        dest="ignore_cistrans",
        action="store_false",
        help=(
            "Compare double-bond cis/trans markers as well as atom chirality."
        ),
    )
    parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=1,
        help=("Atomically refresh Excel after every N evaluated files; use 0 to write only at the end (default: 1)."),
    )
    parser.add_argument(
        "--fail-fast",
        action="store_true",
        help="Stop after the first per-file evaluation error.",
    )
    parser.add_argument(
        "--allow-coverage-mismatch",
        action="store_true",
        help=(
            "Allow missing, extra, duplicate, or ID-less prediction rows. "
            "By default each affected result is an error."
        ),
    )
    return parser


def run_batch(args: argparse.Namespace) -> int:
    print(f"RDKit version: {_runtime_version('rdkit')}", flush=True)

    if not args.gt_path.is_file():
        raise FileNotFoundError(f"Ground-truth JSONL does not exist: {args.gt_path}")
    if not args.results_root.is_dir():
        raise NotADirectoryError(f"Results directory does not exist: {args.results_root}")
    if args.output.suffix.lower() != ".xlsx":
        raise ValueError("--output must end with .xlsx")
    if args.limit is not None and args.limit < 0:
        raise ValueError("--limit must be non-negative")
    if args.max_files is not None and args.max_files < 0:
        raise ValueError("--max-files must be non-negative")
    if args.timeout < 0:
        raise ValueError("--timeout must be non-negative")
    if args.checkpoint_every < 0:
        raise ValueError("--checkpoint-every must be non-negative")

    jobs, skipped_rows, discovery_counts = discover_jobs(
        args.results_root,
        includes=args.include,
        excludes=args.exclude,
    )
    if args.max_files is not None and len(jobs) > args.max_files:
        for job in jobs[args.max_files :]:
            skipped_rows.append(
                {
                    "category": job.category,
                    "model": job.model,
                    "file": job.relative_path,
                    "reason": "Not selected because of --max-files.",
                }
            )
        jobs = jobs[: args.max_files]

    gt_stats = collect_id_stats(args.gt_path)
    gt_contract_errors: list[str] = []
    if gt_stats.read_error_count:
        gt_contract_errors.append(f"{gt_stats.read_error_count} GT read errors")
    if gt_stats.missing_id_count:
        gt_contract_errors.append(
            f"{gt_stats.missing_id_count} GT rows have no string ID"
        )
    if gt_stats.duplicate_count:
        gt_contract_errors.append(f"{gt_stats.duplicate_count} duplicate GT IDs")
    if gt_contract_errors:
        raise ValueError("Invalid ground truth: " + "; ".join(gt_contract_errors))
    evaluated_order = gt_stats.ordered_ids
    if args.limit is not None:
        evaluated_order = evaluated_order[: args.limit]
    evaluated_gt_ids = frozenset(evaluated_order)

    started_at = datetime.now().astimezone().isoformat(timespec="seconds")
    result_rows: list[dict[str, Any]] = []
    initial_info = _run_info(
        args=args,
        discovery_counts=discovery_counts,
        selected_jobs=len(jobs),
        result_rows=result_rows,
        skipped_rows=skipped_rows,
        state="running",
        started_at=started_at,
    )
    write_workbook(args.output, result_rows, skipped_rows, initial_info)

    print(
        f"发现 {discovery_counts['identified_files']} 个可评估文件，本次评估 {len(jobs)} 个，跳过 {len(skipped_rows)} 个。",
        flush=True,
    )
    for index, job in enumerate(jobs, start=1):
        print(
            f"[{index}/{len(jobs)}] {TRACK_LABELS[job.track]} {job.relative_path}",
            flush=True,
        )
        row = evaluate_job(
            job,
            gt_path=args.gt_path,
            full_gt_ids=gt_stats.unique_ids,
            evaluated_gt_ids=evaluated_gt_ids,
            ignore_cistrans=args.ignore_cistrans,
            limit=args.limit,
            timeout_seconds=args.timeout or None,
            allow_coverage_mismatch=args.allow_coverage_mismatch,
        )
        result_rows.append(row)
        accuracy = row["accuracy"]
        accuracy_text = f"{accuracy:.4%}" if accuracy is not None else "N/A"
        print(
            f"    {row['status'].upper()}: {accuracy_text} ({row['elapsed_seconds']:.2f}s)",
            flush=True,
        )

        should_checkpoint = args.checkpoint_every and index % args.checkpoint_every == 0
        if should_checkpoint or row["status"] == "error":
            state = "failed" if args.fail_fast and row["status"] == "error" else "running"
            write_workbook(
                args.output,
                result_rows,
                skipped_rows,
                _run_info(
                    args=args,
                    discovery_counts=discovery_counts,
                    selected_jobs=len(jobs),
                    result_rows=result_rows,
                    skipped_rows=skipped_rows,
                    state=state,
                    started_at=started_at,
                ),
            )
        if args.fail_fast and row["status"] == "error":
            break

    error_count = sum(row["status"] == "error" for row in result_rows)
    completed_all = len(result_rows) == len(jobs)
    if not completed_all:
        final_state = "failed"
    elif error_count:
        final_state = "completed_with_errors"
    else:
        final_state = "completed"
    write_workbook(
        args.output,
        result_rows,
        skipped_rows,
        _run_info(
            args=args,
            discovery_counts=discovery_counts,
            selected_jobs=len(jobs),
            result_rows=result_rows,
            skipped_rows=skipped_rows,
            state=final_state,
            started_at=started_at,
        ),
    )
    print(
        f"评估完成：{len(result_rows)} 个结果，{error_count} 个错误。",
        flush=True,
    )
    print(f"Excel 已保存至: {args.output.resolve()}", flush=True)
    return 1 if error_count else 0


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return run_batch(args)
    except (FileNotFoundError, NotADirectoryError, RuntimeError, ValueError) as exc:
        parser.error(str(exc))
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
